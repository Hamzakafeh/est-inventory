"""
EST Inventory System - Full Read/Write
Alestesharia Animal Nutrition
"""

import os
import sys
import json
import re
import warnings
import threading
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, jsonify, request, session, redirect, url_for

warnings.filterwarnings('ignore')

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.")
    sys.exit(1)

app = Flask(__name__, static_folder='static')
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.secret_key = 'EST-IMS-SecretKey-2026'

# ── بيانات الدخول ──────────────────────────────────────────────────
USERS = {
    'Mlo5': '192.168.100.1',
    'EST':   'Kafeh',
}
# ───────────────────────────────────────────────────────────────────

from flask import send_from_directory

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(app.static_folder, 'est.ico')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET'])
def login_page():
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def do_login():
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    if USERS.get(username) == password:
        session['logged_in'] = True
        session['username'] = username
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'اسم المستخدم أو كلمة المرور غير صحيحة'}), 401

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

MONTH_ORDER = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4,
    'May': 5, 'June': 6, 'July': 7, 'August': 8,
    'September': 9, 'October': 10, 'November': 11, 'December': 12
}
MONTH_AR = {
    'January': 'يناير', 'February': 'فبراير', 'March': 'مارس', 'April': 'أبريل',
    'May': 'مايو', 'June': 'يونيو', 'July': 'يوليو', 'August': 'أغسطس',
    'September': 'سبتمبر', 'October': 'أكتوبر', 'November': 'نوفمبر', 'December': 'ديسمبر'
}

# ── Column mapping (based on VBA: D=4,E=5,F=6,G=7,I=9,J=10,L=12,M=13) ──
# Sheet columns (1-based like Excel):
COL_DATE     = 1   # A - Date
COL_CATEGORY = 4   # D - Category (merged)
COL_TYPE     = 5   # E - Type (merged)
COL_COLOR    = 6   # F - Color
COL_SIZE     = 7   # G - Size
COL_BASIC    = 9   # I - Basic balance
COL_CURRENT  = 10  # J - Current balance
COL_IN       = 12  # L - IN
COL_OUT      = 13  # M - OUT
DATA_START_ROW = 7 # Data starts at row 7

# Log sheet columns
LOG_COL_TIME     = 1
LOG_COL_TYPE     = 2
LOG_COL_QTY      = 3
LOG_COL_BALANCE  = 4
LOG_COL_COLOR    = 5
LOG_COL_SIZE     = 6
LOG_COL_ITEMTYPE = 7
LOG_COL_CATEGORY = 8

def get_base_path():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(script_dir, '2026'),
        os.path.join(os.path.dirname(script_dir), '2026'),
    ]
    for c in candidates:
        if os.path.isdir(c):
            return c
    return None

def get_structure():
    base = get_base_path()
    if not base:
        return {}
    year = os.path.basename(base)
    result = {year: {}}
    for month in sorted(os.listdir(base), key=lambda m: MONTH_ORDER.get(m, 99)):
        month_path = os.path.join(base, month)
        if not os.path.isdir(month_path):
            continue
        files = {}
        for fname in ['Other+', 'Sacks']:
            fpath = os.path.join(month_path, f'{fname}.xlsm')
            if os.path.exists(fpath):
                files[fname] = fpath
        if files:
            result[year][month] = files
    return result

def read_sheet_data(filepath, sheet_name):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None, []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if sheet_name == 'Log':
            if not rows:
                wb.close(); return None, []
            col_indices = [(ci, str(val).strip()) for ci, val in enumerate(rows[0])
                           if val is not None and str(val).strip()]
            if not col_indices:
                wb.close(); return None, []
            data_rows = []
            for row in rows[1:]:
                if not any(ci < len(row) and row[ci] is not None and str(row[ci]).strip() not in ('','None')
                           for ci,_ in col_indices):
                    continue
                rd = {}
                for ci, col_name in col_indices:
                    val = row[ci] if ci < len(row) else None
                    if isinstance(val, datetime): val = val.strftime('%Y-%m-%d %H:%M:%S')
                    elif val is not None:
                        val = str(val) if not isinstance(val,(int,float)) else val
                        if isinstance(val,str) and val.strip()=='None': val=None
                    rd[col_name] = val
                data_rows.append(rd)
            wb.close()
            return [h for _,h in col_indices], data_rows

        if sheet_name == 'Stocktaking':
            data_rows = []
            for row in rows:
                non_empty = [(ci,val) for ci,val in enumerate(row)
                             if val is not None and str(val).strip() not in ('','None')]
                if not non_empty: continue
                rd = {}
                for ci,val in non_empty:
                    col_label = f'Col {ci+1}'
                    if isinstance(val,datetime): val=val.strftime('%Y-%m-%d')
                    elif val is not None: val=str(val) if not isinstance(val,(int,float)) else val
                    rd[col_label]=val
                data_rows.append(rd)
            if not data_rows:
                wb.close(); return None,[]
            all_cols,seen=[],set()
            for r in data_rows:
                for k in r:
                    if k not in seen: all_cols.append(k); seen.add(k)
            wb.close()
            return all_cols, data_rows

        # Main inventory sheet
        header_idx = None
        for i, row in enumerate(rows):
            rv = [str(v).strip() if v else '' for v in row]
            if 'Date' in rv or 'التاريخ' in rv:
                header_idx = i; break
        if header_idx is None:
            wb.close(); return None,[]

        col_indices = [(ci,str(val).strip()) for ci,val in enumerate(rows[header_idx])
                       if val is not None and str(val).strip()]
        data_rows = []
        for row_idx, row in enumerate(rows[header_idx+1:], start=header_idx+2):
            if not any(ci<len(row) and row[ci] is not None and str(row[ci]).strip() not in ('','None')
                       for ci,_ in col_indices):
                continue
            rd={}
            for ci,col_name in col_indices:
                val=row[ci] if ci<len(row) else None
                if isinstance(val,datetime): val=val.strftime('%Y-%m-%d')
                elif val is not None:
                    val=str(val) if not isinstance(val,(int,float)) else val
                    if isinstance(val,str) and val.strip()=='None': val=None
                rd[col_name]=val
            # Store the actual Excel row number for editing
            rd['__row__'] = row_idx
            data_rows.append(rd)
        wb.close()
        headers = [h for _,h in col_indices]
        return headers, data_rows
    except:
        return None, []

# ═══════════════════════════════════════════════════════════════════
#  WRITE LOGIC  —  Replicates the VBA Worksheet_Change macro exactly
# ═══════════════════════════════════════════════════════════════════

def _get_cell_val(ws, row, col):
    """Get value from a cell, handling merged cells by reading the top-left."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    # If it's part of a merged range, get the master cell value
    for merge in ws.merged_cells.ranges:
        if cell.coordinate in merge:
            master = ws.cell(row=merge.min_row, column=merge.min_col)
            return master.value
    return None

def _find_last_balance(ws, target_row, color_value):
    """
    Walk backwards from target_row-1 looking for a row where Color matches.
    Falls back to Basic Balance of the target row if no prior match found.
    Returns (last_balance, found).
    """
    for i in range(target_row - 1, DATA_START_ROW - 1, -1):
        cell_color = ws.cell(row=i, column=COL_COLOR).value
        if cell_color == color_value:
            balance = ws.cell(row=i, column=COL_CURRENT).value
            try:
                return float(balance or 0), True
            except:
                return 0.0, True
    # No previous row with same color — use Basic Balance of THIS row
    basic = ws.cell(row=target_row, column=COL_BASIC).value
    try:
        return float(basic or 0), False
    except:
        return 0.0, False

def _append_log(ws_log, operation, qty, balance, color, size, item_type, category):
    """Append a row to the Log sheet."""
    lr = 1
    for row in ws_log.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value is not None:
                lr = cell.row
    lr += 1
    ws_log.cell(row=lr, column=LOG_COL_TIME).value     = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_log.cell(row=lr, column=LOG_COL_TYPE).value     = operation
    ws_log.cell(row=lr, column=LOG_COL_QTY).value      = qty
    ws_log.cell(row=lr, column=LOG_COL_BALANCE).value  = balance
    ws_log.cell(row=lr, column=LOG_COL_COLOR).value    = color
    ws_log.cell(row=lr, column=LOG_COL_SIZE).value     = size
    ws_log.cell(row=lr, column=LOG_COL_ITEMTYPE).value = item_type
    ws_log.cell(row=lr, column=LOG_COL_CATEGORY).value = category

# ── Flask routes ────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════
#  Stocktaking auto-recalculation
#  Parses LOOKUP(2,1/(Sheet!Col:Col=Val),Sheet!J:J) formulas and
#  computes the result in Python, then writes it back to the cell.
# ══════════════════════════════════════════════════════════════════
_LOOKUP_RE = re.compile(
    r'LOOKUP\s*\(\s*2\s*,\s*1\s*/\s*\(\s*(\w+)!\s*([A-Z]+)\s*:\s*\2\s*=\s*(.*?)\s*\)\s*,\s*\1!\s*([A-Z]+)\s*:\s*\4\s*\)',
    re.IGNORECASE
)

def _col_letter_to_idx(letter):
    """'A'->0, 'B'->1, ... 'J'->9  (0-based for row tuple indexing)"""
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1  # 0-based

def _recalc_stocktaking(wb):
    """
    Find the Stocktaking sheet in a workbook, parse every LOOKUP formula,
    evaluate it by scanning the referenced sheet, and write the computed
    value back.  Modifies wb in-place (caller must save).
    """
    if 'Stocktaking' not in wb.sheetnames:
        return

    ws_st = wb['Stocktaking']

    # Cache sheet rows for performance
    _sheet_rows_cache = {}
    def _get_rows(sheet_name):
        if sheet_name not in _sheet_rows_cache:
            if sheet_name in wb.sheetnames:
                _sheet_rows_cache[sheet_name] = list(
                    wb[sheet_name].iter_rows(min_row=DATA_START_ROW, values_only=True)
                )
            else:
                _sheet_rows_cache[sheet_name] = []
        return _sheet_rows_cache[sheet_name]

    for row in ws_st.iter_rows():
        for cell in row:
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                continue
            m = _LOOKUP_RE.search(cell.value)
            if not m:
                continue

            ref_sheet   = m.group(1)
            filter_col  = _col_letter_to_idx(m.group(2))   # e.g. F -> 5
            raw_val     = m.group(3).strip().strip('"')     # filter value
            result_col  = _col_letter_to_idx(m.group(4))   # e.g. J -> 9

            # Auto-cast numeric filter value
            try:
                filter_val = float(raw_val) if '.' in raw_val else int(raw_val)
            except ValueError:
                filter_val = raw_val  # keep as string

            # LOOKUP(2,1/(...)) = last matching row (walk forward, keep updating)
            last_result = None
            for data_row in _get_rows(ref_sheet):
                try:
                    cell_filter = data_row[filter_col]
                    # Normalize comparison: strip strings, cast numbers
                    if isinstance(filter_val, str):
                        match = (cell_filter is not None and
                                 str(cell_filter).strip() == filter_val)
                    else:
                        try:
                            match = float(cell_filter) == filter_val
                        except (TypeError, ValueError):
                            match = False

                    if match:
                        v = data_row[result_col]
                        if v is not None:
                            try:
                                last_result = float(v)
                            except (TypeError, ValueError):
                                last_result = v
                except IndexError:
                    continue

            # Write computed value (keep formula string intact, just overwrite value)
            cell.value = last_result if last_result is not None else cell.value

@app.route('/')
@login_required
def index():
    return render_template('index.html', structure=get_structure(),
                           base_path=get_base_path() or 'Not found', month_ar=MONTH_AR)

@app.route('/api/structure')
@login_required
def api_structure():
    return jsonify(get_structure())

@app.route('/api/sheets')
@login_required
def api_sheets():
    filepath = request.args.get('path')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames; wb.close()
        return jsonify({'sheets': sheets})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data')
@login_required
def api_data():
    filepath = request.args.get('path')
    sheet    = request.args.get('sheet','')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    headers, rows = read_sheet_data(filepath, sheet)
    if headers is None:
        return jsonify({'headers':[],'rows':[],'count':0})
    return jsonify({'headers': headers, 'rows': rows, 'count': len(rows)})


# ══════════════════════════════════════════════════════════════════
#  NEW: /api/transaction  —  IN or OUT on the main inventory sheet
#
#  Body (JSON):
#    filepath  : full path to the .xlsm file
#    sheet     : sheet name (e.g. "Sheet1")
#    row       : Excel row number (integer, from __row__ field)
#    operation : "IN" or "OUT"
#    qty       : positive number
# ══════════════════════════════════════════════════════════════════
@app.route('/api/transaction', methods=['POST'])
@login_required
def api_transaction():
    data = request.get_json(silent=True) or {}
    filepath  = data.get('filepath', '')
    sheet     = data.get('sheet', '')
    row       = data.get('row')
    operation = data.get('operation', '').upper()
    qty_raw   = data.get('qty')

    # Validate
    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    if operation not in ('IN', 'OUT'):
        return jsonify({'success': False, 'error': 'operation must be IN or OUT'}), 400
    try:
        row = int(row)
        qty = float(qty_raw)
        if qty < 0:
            raise ValueError()
    except:
        return jsonify({'success': False, 'error': 'Invalid row or qty'}), 400

    try:
        # Load workbook keeping VBA (macros stay intact)
        wb = openpyxl.load_workbook(filepath, keep_vba=True)

        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404

        ws      = wb[sheet]
        ws_log  = wb['Log'] if 'Log' in wb.sheetnames else None

        # Read context values (mirrors VBA variable reads)
        color_value    = _get_cell_val(ws, row, COL_COLOR)
        size_value     = _get_cell_val(ws, row, COL_SIZE)
        type_value     = _get_cell_val(ws, row, COL_TYPE)
        category_value = _get_cell_val(ws, row, COL_CATEGORY)
        basic_balance  = ws.cell(row=row, column=COL_BASIC).value or 0

        # Require Color to be set before any transaction
        if not color_value or str(color_value).strip() in ('', 'None', 'null'):
            wb.close()
            return jsonify({'success': False,
                            'error': 'يجب تحديد اللون (Color) أولاً قبل إجراء أي عملية'}), 400

        # Find last balance for same colour (VBA loop)
        last_balance, found = _find_last_balance(ws, row, color_value)
        if not found:
            try:
                last_balance = float(basic_balance)
            except:
                last_balance = 0.0

        # Compute new balance
        if operation == 'IN':
            new_balance = last_balance + qty
        else:
            new_balance = last_balance - qty

        # Write current balance (column J)
        ws.cell(row=row, column=COL_CURRENT).value = new_balance

        # Append to Log sheet
        if ws_log:
            _append_log(ws_log, operation, qty, new_balance,
                        color_value, size_value, type_value, category_value)

        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()

        return jsonify({
            'success':     True,
            'new_balance': new_balance,
            'operation':   operation,
            'qty':         qty,
            'color':       color_value,
            'size':        size_value,
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  NEW: /api/update_cell  —  Edit any plain cell directly
#
#  Body (JSON):
#    filepath : full path to the .xlsm file
#    sheet    : sheet name
#    row      : Excel row number
#    col_name : column header name (e.g. "Date", "Color")
#    value    : new value (string; numbers auto-cast)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/update_cell', methods=['POST'])
@login_required
def api_update_cell():
    data      = request.get_json(silent=True) or {}
    filepath  = data.get('filepath', '')
    sheet     = data.get('sheet', '')
    row       = data.get('row')
    col_name  = data.get('col_name', '')
    value     = data.get('value', '')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    try:
        row = int(row)
    except:
        return jsonify({'success': False, 'error': 'Invalid row'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404
        ws = wb[sheet]

        # Find column index by scanning header row
        header_row_idx = None
        for i, row_cells in enumerate(ws.iter_rows(values_only=True), start=1):
            rv = [str(v).strip() if v else '' for v in row_cells]
            if 'Date' in rv or 'التاريخ' in rv:
                header_row_idx = i; break

        if header_row_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': 'Header row not found'}), 400

        col_idx = None
        for ci, cell in enumerate(ws[header_row_idx], start=1):
            if cell.value and str(cell.value).strip() == col_name:
                col_idx = ci; break

        if col_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': f'Column "{col_name}" not found'}), 400

        # Auto-cast numeric values
        cast_value = value
        try:
            if '.' in str(value):
                cast_value = float(value)
            else:
                cast_value = int(value)
        except:
            cast_value = value if value != '' else None

        ws.cell(row=row, column=col_idx).value = cast_value
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()

        return jsonify({'success': True, 'row': row, 'col': col_name, 'value': cast_value})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  /api/color_balance  —  Get last current balance for a Color
#
#  Query params: path, sheet, color, before_row (optional)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/color_balance')
@login_required
def api_color_balance():
    filepath   = request.args.get('path', '')
    sheet      = request.args.get('sheet', '')
    color      = request.args.get('color', '')
    before_row = request.args.get('before_row', None)

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    if not color or color.strip().lower() in ('', 'null', 'none'):
        return jsonify({'balance': None, 'found': False})

    try:
        before_row = int(before_row) if before_row else None
    except:
        before_row = None

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'balance': None, 'found': False})
        ws = wb[sheet]

        last_balance = None
        last_row     = None
        rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))
        for idx, row_vals in enumerate(rows):
            actual_row = DATA_START_ROW + idx
            if before_row and actual_row >= before_row:
                break
            # Color is COL_COLOR (6) → index 5 in 0-based row_vals
            # but row_vals starts from col 1, so col 6 → index 5
            try:
                cell_color   = row_vals[COL_COLOR - 1]
                cell_current = row_vals[COL_CURRENT - 1]
            except IndexError:
                continue
            if cell_color and str(cell_color).strip() == color.strip():
                try:
                    val = float(cell_current) if cell_current is not None else None
                    if val is not None:
                        last_balance = val
                        last_row     = actual_row
                except:
                    pass

        wb.close()
        return jsonify({'balance': last_balance, 'found': last_balance is not None, 'row': last_row})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  /api/set_opening_balance  —  Write Basic + Current in one call
#
#  Body: filepath, sheet, row, balance (number)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/set_opening_balance', methods=['POST'])
@login_required
def api_set_opening_balance():
    data     = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet    = data.get('sheet', '')
    row      = data.get('row')
    balance  = data.get('balance')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    try:
        row     = int(row)
        balance = float(balance)
    except:
        return jsonify({'success': False, 'error': 'Invalid row or balance'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        ws = wb[sheet]
        ws.cell(row=row, column=COL_BASIC).value   = balance
        ws.cell(row=row, column=COL_CURRENT).value = balance
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()
        return jsonify({'success': True, 'balance': balance})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  NEW: /api/add_row  —  Add a new item row to the sheet
#
#  Body (JSON):
#    filepath : full path to the .xlsm file
#    sheet    : sheet name
#    fields   : dict of { col_name: value, ... }
# ══════════════════════════════════════════════════════════════════
def _parse_dv_formula(formula1):
    """Parse a Data Validation formula1 string into a clean list of values."""
    if not formula1:
        return []
    s = formula1.strip().strip('"')
    items = [v.strip() for v in s.split(',') if v.strip() and v.strip().lower() not in ('null', 'none', '')]
    return items

def _col_letter_to_index(letter):
    """Convert Excel column letter(s) to 1-based index. E.g. 'F' -> 6."""
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result

@app.route('/api/options')
@login_required
def api_options():
    """Read Data Validation lists from the Excel sheet for Color, Type, Size, Category."""
    filepath = request.args.get('path', '')
    sheet    = request.args.get('sheet', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    try:
        # Must open with read_only=False to access data_validations
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'colors': [], 'types': [], 'sizes': [], 'categories': []})
        ws = wb[sheet]

        # Find which column letter maps to Color(F=6), Type(E=5), Size(G=7), Category(D=4)
        # by reading the header row (row 6) dynamically
        header_col_map = {}  # col_index (1-based) -> header name
        for ci, cell in enumerate(ws[6], start=1):
            if cell.value and str(cell.value).strip():
                header_col_map[ci] = str(cell.value).strip()

        # Build reverse: header name -> col_index
        name_to_col = {v: k for k, v in header_col_map.items()}

        col_color    = name_to_col.get('Color',    COL_COLOR)
        col_type     = name_to_col.get('Type',     COL_TYPE)
        col_size     = name_to_col.get('Size',     COL_SIZE)
        col_category = name_to_col.get('Category', COL_CATEGORY)

        options = {'colors': [], 'types': [], 'sizes': [], 'categories': []}
        col_target = {
            col_color:    'colors',
            col_type:     'types',
            col_size:     'sizes',
            col_category: 'categories',
        }

        for dv in ws.data_validations.dataValidation:
            if dv.type != 'list' or not dv.formula1:
                continue
            # Get the column index from sqref (e.g. "F7:F1048576" -> col F -> 6)
            try:
                first_ref = str(dv.sqref).split()[0]   # take first range if multiple
                col_letters = ''.join(c for c in first_ref.split(':')[0] if c.isalpha())
                ci = _col_letter_to_index(col_letters)
            except Exception:
                continue

            key = col_target.get(ci)
            if key:
                vals = _parse_dv_formula(dv.formula1)
                options[key] = vals  # last DV for that col wins

        wb.close()
        return jsonify(options)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clear_row', methods=['POST'])
@login_required
def api_clear_row():
    """Clear all data cells in a given Excel row (does NOT delete the row itself)."""
    data     = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet    = data.get('sheet', '')
    row      = data.get('row')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    try:
        row = int(row)
    except:
        return jsonify({'success': False, 'error': 'Invalid row'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        ws = wb[sheet]
        # Clear columns A–M (1–13) — keeps the row intact for VBA structure
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.value = None
        _recalc_stocktaking(wb)
        wb.save(filepath)
        wb.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/add_row', methods=['POST'])
@login_required
def api_add_row():
    data     = request.get_json(silent=True) or {}
    filepath = data.get('filepath', '')
    sheet    = data.get('sheet', '')
    fields   = data.get('fields', {})

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    if not fields:
        return jsonify({'success': False, 'error': 'No fields provided'}), 400

    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': f'Sheet "{sheet}" not found'}), 404
        ws = wb[sheet]

        # Find header row
        header_row_idx = None
        for i, row_cells in enumerate(ws.iter_rows(values_only=True), start=1):
            rv = [str(v).strip() if v else '' for v in row_cells]
            if 'Date' in rv or 'التاريخ' in rv:
                header_row_idx = i; break

        if header_row_idx is None:
            wb.close()
            return jsonify({'success': False, 'error': 'Header row not found'}), 400

        # Build col_name -> col_index map
        col_map = {}
        for ci, cell in enumerate(ws[header_row_idx], start=1):
            if cell.value:
                col_map[str(cell.value).strip()] = ci

        # Find next empty data row (search from DATA_START_ROW downward)
        new_row = DATA_START_ROW
        for r in range(DATA_START_ROW, ws.max_row + 2):
            row_empty = True
            for ci in range(1, 14):
                if ws.cell(row=r, column=ci).value not in (None, ''):
                    row_empty = False
                    break
            if row_empty:
                new_row = r
                break

        # Write fields
        for col_name, value in fields.items():
            if col_name in col_map:
                # Auto-cast
                cast_value = value
                try:
                    if '.' in str(value):
                        cast_value = float(value)
                    else:
                        cast_value = int(value)
                except:
                    cast_value = value if value != '' else None
                ws.cell(row=new_row, column=col_map[col_name]).value = cast_value

        # Set Date if not provided
        if 'Date' not in fields and 'التاريخ' not in fields:
            date_col = col_map.get('Date') or col_map.get('التاريخ')
            if date_col:
                ws.cell(row=new_row, column=date_col).value = datetime.now().strftime('%Y-%m-%d')

        wb.save(filepath)
        wb.close()
        return jsonify({'success': True, 'new_row': new_row})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    try:
        from flaskwebgui import FlaskUI
        ui = FlaskUI(
            app=app,
            server='flask',
            width=1280,
            height=800,
            port=3049,
            fullscreen=False,
        )
        ui.run()
    except ImportError:
        import webbrowser, threading, time
        def _open():
            time.sleep(1.2)
            webbrowser.open('http://127.0.0.1:3049')
        threading.Thread(target=_open, daemon=True).start()
        app.run(host='127.0.0.1', port=3049, debug=False)
