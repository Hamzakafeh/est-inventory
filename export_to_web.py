"""
EST Inventory System — Export to Web
=====================================
شغّل هذا السكريبت بعد أي تحديث على ملفات Excel.
يحول كل البيانات إلى data.json جاهز للرفع على GitHub.

الاستخدام:
    python export_to_web.py

الناتج:
    data.json  (في نفس مجلد السكريبت)
"""

import os
import json
import warnings
from datetime import datetime

warnings.filterwarnings('ignore')

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl غير مثبت. شغّل: pip install openpyxl")
    input("اضغط Enter للخروج...")
    exit(1)

# ── إعدادات المسار ──────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# يبحث عن مجلد 2026 في نفس المجلد أو المجلد الأعلى
CANDIDATES = [
    os.path.join(SCRIPT_DIR, '2026'),
    os.path.join(SCRIPT_DIR, 'data', '2026'),
    os.path.join(os.path.dirname(SCRIPT_DIR), 'data', '2026'),
]

BASE_PATH = None
for c in CANDIDATES:
    if os.path.isdir(c):
        BASE_PATH = c
        break

MONTH_ORDER = {
    'January':1,'February':2,'March':3,'April':4,
    'May':5,'June':6,'July':7,'August':8,
    'September':9,'October':10,'November':11,'December':12
}

# ── قراءة الشيتات ──────────────────────────────────────────────────

def read_log_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    col_indices = [(ci, str(val).strip()) for ci, val in enumerate(rows[0])
                   if val is not None and str(val).strip()]
    if not col_indices:
        return [], []
    data_rows = []
    for row in rows[1:]:
        if not any(ci < len(row) and row[ci] is not None and
                   str(row[ci]).strip() not in ('', 'None') for ci, _ in col_indices):
            continue
        rd = {}
        for ci, cn in col_indices:
            val = row[ci] if ci < len(row) else None
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d %H:%M:%S')
            elif val is not None:
                val = str(val) if not isinstance(val, (int, float)) else val
                if isinstance(val, str) and val.strip() == 'None':
                    val = None
            rd[cn] = val
        data_rows.append(rd)
    return [h for _, h in col_indices], data_rows


def read_main_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        rv = [str(v).strip() if v else '' for v in row]
        if 'Date' in rv or 'التاريخ' in rv:
            header_idx = i
            break
    if header_idx is None:
        return [], []
    col_indices = [(ci, str(val).strip()) for ci, val in enumerate(rows[header_idx])
                   if val is not None and str(val).strip()]
    data_rows = []
    for row in rows[header_idx + 1:]:
        if not any(ci < len(row) and row[ci] is not None and
                   str(row[ci]).strip() not in ('', 'None') for ci, _ in col_indices):
            continue
        rd = {}
        for ci, cn in col_indices:
            val = row[ci] if ci < len(row) else None
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d')
            elif val is not None:
                val = str(val) if not isinstance(val, (int, float)) else val
                if isinstance(val, str) and val.strip() == 'None':
                    val = None
            rd[cn] = val
        data_rows.append(rd)
    return [h for _, h in col_indices], data_rows


def read_stocktaking(ws):
    rows = list(ws.iter_rows(values_only=True))
    data_rows = []
    for row in rows:
        non_empty = [(ci, val) for ci, val in enumerate(row)
                     if val is not None and str(val).strip() not in ('', 'None')]
        if not non_empty:
            continue
        rd = {}
        for ci, val in non_empty:
            col_label = f'Col {ci + 1}'
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d')
            elif val is not None:
                val = str(val) if not isinstance(val, (int, float)) else val
            rd[col_label] = val
        data_rows.append(rd)
    if not data_rows:
        return [], []
    all_cols, seen = [], set()
    for r in data_rows:
        for k in r:
            if k not in seen:
                all_cols.append(k)
                seen.add(k)
    return all_cols, data_rows


# ── التصدير الرئيسي ────────────────────────────────────────────────

def export():
    if not BASE_PATH:
        print("❌ لم يتم العثور على مجلد 2026!")
        print("   تأكد أن السكريبت في نفس مجلد البرنامج.")
        input("اضغط Enter للخروج...")
        return

    print(f"📂 مجلد البيانات: {BASE_PATH}")
    print("⏳ جاري التصدير...\n")

    result = {}
    month_folders = sorted(
        [m for m in os.listdir(BASE_PATH) if os.path.isdir(os.path.join(BASE_PATH, m))],
        key=lambda m: MONTH_ORDER.get(m.split('-', 1)[1] if '-' in m else m, 99)
    )

    total_sheets = 0
    total_rows   = 0

    for month_folder in month_folders:
        month_name = month_folder.split('-', 1)[1] if '-' in month_folder else month_folder
        month_path = os.path.join(BASE_PATH, month_folder)
        result[month_name] = {}

        for fname in ['Sacks', 'Other+']:
            fpath = os.path.join(month_path, f'{fname}.xlsm')
            if not os.path.exists(fpath):
                continue

            result[month_name][fname] = {}
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                for sh in wb.sheetnames:
                    ws = wb[sh]
                    if sh == 'Log':
                        headers, rows = read_log_sheet(ws)
                    elif sh == 'Stocktaking':
                        headers, rows = read_stocktaking(ws)
                    else:
                        headers, rows = read_main_sheet(ws)

                    result[month_name][fname][sh] = {
                        'headers': headers,
                        'rows':    rows,
                        'count':   len(rows)
                    }
                    total_sheets += 1
                    total_rows   += len(rows)

                wb.close()
                print(f"  ✅ {month_name} / {fname}")
            except Exception as e:
                print(f"  ⚠️  {month_name} / {fname} — خطأ: {e}")

    output = {
        'generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'data': result
    }

    out_path = os.path.join(SCRIPT_DIR, 'data.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

    size_kb = os.path.getsize(out_path) / 1024
    print(f"\n✅ تم التصدير بنجاح!")
    print(f"   📄 data.json  ({size_kb:.1f} KB)")
    print(f"   📊 {total_sheets} شيت  |  {total_rows} صف")
    print(f"\n📌 الخطوة التالية:")
    print(f"   ارفع ملف data.json على GitHub Repository")
    print(f"   (راجع ملف INSTRUCTIONS.md للتفاصيل)")
    input("\nاضغط Enter للخروج...")


if __name__ == '__main__':
    export()